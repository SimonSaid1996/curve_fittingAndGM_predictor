import numpy as np
from scipy.optimize import curve_fit         #later used for curveFit
import openpyxl                              #for excel reading
import math
from GM_predictor import GM_predictor

'''author ziran cao
a simple predicting and testing program which will enable users to input some data, then use the existing data to predict
how the later series of data will look like

'''
def getCorrectionRate1(predicting_data,actual_data):
    for i in range(len(predicting_data)):
        predictElement = predicting_data[i]
        actualElement = actual_data[i]
        errorRate = (actualElement - predictElement)/actualElement
        print('real value ',actualElement,'predicting value',predictElement,'error rate',errorRate)

def curvFitFunc(x, a, b, c,e):   #适当情况下，项数越少，精度越高，但是要有后面的阈值作为调节
    return a*np.cos(b*x*np.pi/180+c*np.pi/180)+e  #+d*x

def NoneLinear_leastSquare(date_train, data_train,date_test,data_test):
    popt, pcov = curve_fit(curvFitFunc, date_train, data_train)  # 非线性最小二乘法，用的是实际值与预测值误差的平方
    a = popt[0];
    b = popt[1];
    c = popt[2];
    e = popt[3];# getting the coeffecient numbers for the function
    datatest_predict = curvFitFunc(date_test, a, b, c,  e)  #d,
    return datatest_predict

def processLists(date_list, data_list, col_range,date_appended):
    for i in range( len(col_range)):  # col in col_range1
        for j in range( len( col_range[i])):#for cell in col:
            if i == 0 and j != 0 and date_appended == False:  #don't append if date_store_list length is not empty
                date_list.append( j ) # appending all the date vlaues in the date_store_list, by month
            if i== 1 and j != 0:
                data_list.append( col_range[i][j].value ) # appending all the 支出 values in the data_store_list, not including the title
    return date_list,data_list

def process_lists(spend1, spend2, gain1, gain2):
    #assuming all four lists have the same length
    original_DKYE_num = 2425302846.89; original_GJYE_num = 2881947995.2
    DKYE_list =[original_DKYE_num]; GJYE_list=[original_GJYE_num]
    for i in range( len(spend1) ):  #adding up values and appending in the DKYE and GJYE lists
        if i == 0:  #first element, need to add the initial value
            DKYE_temp =  DKYE_list[i]
            GJYE_temp =  GJYE_list[i]
            DKYE_list[i] = DKYE_temp + spend1[i] - gain2[i]
            GJYE_list[i] = GJYE_temp + gain1[i] - spend2[i]   #recalculate the first term
        else:    #adding up the later terms
            DKYE = DKYE_list[i-1] + spend1[i] - gain2[i]
            GJYE = GJYE_list[i-1] + gain1[i] - spend2[i]
            DKYE_list.append(DKYE)
            GJYE_list.append(GJYE)
    return DKYE_list, GJYE_list

# Press the green button in the gutter to run the script.
if __name__ == '__main__':           #to think of the improvement,could go thinking about normalization, neuronN or activation function
    wb_s1 = openpyxl.load_workbook('支出1.xlsx')
    ws_s1 = wb_s1.active   #choosing the active sheets
    col_range_s1 = ws_s1['B:C']  #B:C
    wb_s2 = openpyxl.load_workbook('支出2.xlsx')
    ws_s2 = wb_s2.active  # choosing the active sheets
    col_range_s2 = ws_s2['B:C']  # B:C
    wb_g1 = openpyxl.load_workbook('收入1.xlsx')
    ws_g1 = wb_g1.active  # choosing the active sheets
    col_range_g1 = ws_g1['B:C']  # B:C
    wb_g2 = openpyxl.load_workbook('收入2.xlsx')
    ws_g2 = wb_g2.active   #choosing the active sheets
    col_range_g2 = ws_g2['B:C']  #B:C

    date_stroe_list =[]   #to store the date list
    spend1_list = []  #to store the data number list
    spend2_list =[]
    gain1_list = []
    gain2_list = []

    date_stroe_list, spend1_list = processLists(date_stroe_list, spend1_list, col_range_s1, date_appended= False)
    date_stroe_list, spend2_list = processLists(date_stroe_list, spend2_list, col_range_s2, date_appended= True)
    date_stroe_list, gain1_list = processLists(date_stroe_list, gain1_list, col_range_g1, date_appended=True)
    date_stroe_list, gain2_list = processLists(date_stroe_list, gain2_list, col_range_g2, date_appended=True)


    '''#GM_predictor, need to use spend1's data to explain what happened, calculation by month
    #one year situation prediction
    spend1_train = np.array(spend1_list[:12])
    spend1_test = np.array(spend1_list[12:24])   #only for testing GM_predictor
    spend2_train = np.array(spend2_list[:12])
    spend2_test = np.array(spend2_list[12:24])
    gain1_train = np.array(gain1_list[:12])
    gain1_test = np.array(gain1_list[12:24])
    gain2_train = np.array(gain2_list[:12])
    gain2_test = np.array(gain2_list[12:24])
    date_train = np.array(date_stroe_list[:12])
    date_test = np.array(date_stroe_list[12:24])

    #8 month with clear trend situation
    #spend1_train = np.array(spend1_list[25:37])  
    #spend1_test = np.array(spend1_list[37:45])  
    #spend2_train = np.array(spend2_list[25:37])
    #spend2_test = np.array(spend2_list[37:45])
    #gain1_train = np.array(gain1_list[25:37])
    #gain1_test = np.array(gain1_list[37:45])
    #gain2_train = np.array(gain2_list[25:37])
    #gain2_test = np.array(gain2_list[37:45])
    #date_train = np.array(date_stroe_list[25:37])
    #date_test = np.array(date_stroe_list[37:45])


    #1 year with not clear trend situation
    #spend1_train = np.array(spend1_list[12:24])#np.array(spend1_list[:60])
    #spend1_test = np.array(spend1_list[24:36]) #np.array(spend1_list[60:])   #only for testing GM_predictor
    #spend2_train = np.array(spend2_list[12:24])
    #spend2_test = np.array(spend2_list[24:36])
    #gain1_train = np.array(gain1_list[12:24])
    #gain1_test = np.array(gain1_list[24:36])
    #gain2_train = np.array(gain2_list[12:24])
    #gain2_test = np.array(gain2_list[24:36])
    #date_train = np.array(date_stroe_list[12:24])
    #date_test = np.array(date_stroe_list[24:36])

    myGM_predictor = GM_predictor(spend1_train)
    spend1_pred = myGM_predictor.GM11_predict( 12) #(12)  #predicting the data after one year or 3 months
    myGM_predictor = GM_predictor(spend2_train)
    spend2_pred = myGM_predictor.GM11_predict( 12) #(12)
    myGM_predictor = GM_predictor(gain1_train)
    gain1_pred = myGM_predictor.GM11_predict( 12) #(12)
    myGM_predictor = GM_predictor(gain2_train)
    gain2_pred = myGM_predictor.GM11_predict( 12) #(12)
    DKYE_real, GJYE_real = process_lists(spend1_test, spend2_test, gain1_test, gain2_test)  # getting the real values
    DKYE_pred, GJYE_pred = process_lists(spend1_pred, spend2_pred, gain1_pred, gain2_pred)
    print('below is DKYE predictions')
    getCorrectionRate1(DKYE_pred, DKYE_real)
    print('below is GJYE predictions')
    getCorrectionRate1(GJYE_pred,GJYE_real )
    '''
    
    #'''
    #divide the data len into train and test, and put them in to array for later calculations, 60 units for train and 38 units for test
    #curve_fitting,depending on how the data flucturated, need to adjust the curve functions
    spend1_train = np.array(spend1_list[:60])
    spend1_test = np.array(spend1_list[60:])
    spend2_train = np.array(spend2_list[:60])
    spend2_test = np.array(spend2_list[60:])
    gain1_train = np.array(gain1_list[:60])
    gain1_test = np.array(gain1_list[60:])
    gain2_train = np.array(gain2_list[:60])
    gain2_test = np.array(gain2_list[60:])
    date_train = np.array(date_stroe_list[:60])
    date_test = np.array(date_stroe_list[60:])
    spend1_pred = NoneLinear_leastSquare(date_train, spend1_train, date_test, spend1_test)  #most times works, some extreme values for 支出1，收
    spend2_pred = NoneLinear_leastSquare(date_train, spend2_train, date_test, spend2_test)
    gain1_pred = NoneLinear_leastSquare(date_train, gain1_train, date_test, gain1_test)
    gain2_pred = NoneLinear_leastSquare(date_train, gain2_train, date_test, gain2_test)
    #getting the result values by formula
    DKYE_real, GJYE_real = process_lists(spend1_test, spend2_test, gain1_test, gain2_test)  #getting the real values
    DKYE_pred, GJYE_pred = process_lists(spend1_pred, spend2_pred, gain1_pred, gain2_pred)
    print('below testing the DKYE data\n')
    getCorrectionRate1(DKYE_pred, DKYE_real)
    print('below testing the GJYE data\n')
    getCorrectionRate1(GJYE_pred, GJYE_real)
    #'''

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
